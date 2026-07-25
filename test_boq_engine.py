import os
import io
import openpyxl
from openpyxl.styles import PatternFill
from rag_engine import get_chroma_client, get_or_create_collection
from exporter import process_boq_excel

def test_boq_flow():
    print("--- Testing BOQ / Excel Auto-Fill Engine ---")
    
    # 1. Create a dummy tender BOQ Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tender Compliance BOQ"
    
    headers = ["Sl No", "Item Description / Specifications", "Qty", "Unit"]
    ws.append(headers)
    
    sample_items = [
        [1, "Rack Server 2U: Dual Intel Xeon 6330 28C, 256GB ECC DDR4 RAM, 2x 1.92TB NVMe SSD", 4, "Nos"],
        [2, "Managed L3 Core Switch: 48-Port 10G SFP+ with redundant hot-swap power supplies", 2, "Nos"],
        [3, "Next-Gen Enterprise Firewall: 10 Gbps Threat Protection throughput with IPS/SSL Inspection", 1, "Nos"]
    ]
    for row in sample_items:
        ws.append(row)
        
    in_stream = io.BytesIO()
    wb.save(in_stream)
    excel_bytes = in_stream.getvalue()
    print(f"Created sample input BOQ Excel file ({len(excel_bytes)} bytes)")
    
    # 2. Get ChromaDB collection
    client = get_chroma_client()
    collection = get_or_create_collection(client)
    
    # 3. Process BOQ in Demo Mode
    output_bytes, processed_rows = process_boq_excel(
        excel_bytes,
        collection,
        api_key="demo_key",
        demo_mode=True
    )
    
    print(f"Processed {len(processed_rows)} BOQ items!")
    for item in processed_rows:
        print(f"  Row {item['row_num']}: [{item['compliance']}] {item['item'][:50]}...")
        print(f"    Spec: {item['response'][:70]}...")
        
    # 4. Validate output Excel structure
    out_wb = openpyxl.load_workbook(io.BytesIO(output_bytes))
    out_ws = out_wb.active
    
    print("\nOutput Excel Columns:")
    col_headers = [out_ws.cell(row=1, column=c).value for c in range(1, out_ws.max_column + 1)]
    print(" ", col_headers)
    
    assert "AI Compliance Status" in col_headers, "AI Compliance Status column missing!"
    assert "AI Technical Specification & Response" in col_headers, "AI Technical Spec column missing!"
    assert "AI Remarks & Sourced Context" in col_headers, "AI Remarks column missing!"
    
    print("[PASS] BOQ / Excel Auto-Fill Engine Test PASSED cleanly!")

if __name__ == "__main__":
    test_boq_flow()
